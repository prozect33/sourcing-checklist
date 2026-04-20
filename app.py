import streamlit as st
import json
import os
import pandas as pd
import datetime
import uuid
import re
import tempfile
import csv
from dataclasses import dataclass
import openpyxl

from html.parser import HTMLParser
from ad_analysis_tab import render_ad_analysis_tab
from supabase import create_client, Client
from typing import Dict, List, Tuple

st.set_page_config(page_title="간단 마진 계산기", layout="wide")

st.markdown("""
    <style>
     [data-testid="stSidebarHeader"] { display: none !important; }
     [data-testid="stSidebarContent"] { padding-top: 15px !important; }
     [data-testid="stHeading"] { margin-bottom: 15px !important; }
     [data-testid="stNumberInput"] button { display: none !important; }
    </style>
""", unsafe_allow_html=True)

def format_number(val):
    if val is None:
        return ""
    return f"{int(val):,}" if float(val).is_integer() else f"{val:,.2f}"

def reset_inputs():
    # 탭1 리셋
    st.session_state["sell_price_raw"] = ""
    st.session_state["unit_yuan"] = ""
    st.session_state["unit_won"] = ""
    st.session_state["qty_raw"] = ""
    st.session_state["show_result"] = False
    
    # 탭2 일일 정산 리셋
    if "total_sales_qty" in st.session_state: st.session_state["total_sales_qty"] = 0
    if "total_revenue" in st.session_state: st.session_state["total_revenue"] = 0
    if "ad_sales_qty" in st.session_state: st.session_state["ad_sales_qty"] = 0
    if "ad_revenue" in st.session_state: st.session_state["ad_revenue"] = 0
    if "ad_cost" in st.session_state: st.session_state["ad_cost"] = 0
    if "coupon_unit" in st.session_state: st.session_state["coupon_unit"] = 0
    if "product_select_daily" in st.session_state:
       st.session_state["product_select_daily"] = "상품을 선택해주세요"

def load_supabase_credentials():
    try:
        with open("credentials.json", "r") as f:
            creds = json.load(f)
            return creds["SUPABASE_URL"], creds["SUPABASE_KEY"]
    except FileNotFoundError:
        st.error("오류: 'credentials.json' 파일을 찾을 수 없습니다.\n파일을 생성하고 Supabase 키를 입력해주세요.")
        st.stop()
    except json.JSONDecodeError:
        st.error("오류: 'credentials.json' 파일의 형식이 잘못되었습니다. JSON 형식을 확인해주세요.")
        st.stop()
    except KeyError:
        st.error("오류: 'credentials.json' 파일에 'SUPABASE_URL' 또는 'SUPABASE_KEY'가 없습니다.")
        st.stop()

# ← 사이드바 시작
try:
    SUPABASE_URL, SUPABASE_KEY = load_supabase_credentials()
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except Exception as e:
    st.error(f"Supabase 클라이언트 초기화 중 오류가 발생했습니다: {e}")
    st.stop()

def _fetch_all_rows(table: str, select: str, *, batch_size: int = 1000) -> List[dict]:
    rows: List[dict] = []
    start = 0
    while True:
        q = supabase.table(table).select(select)
        try:
            resp = q.range(start, start + batch_size - 1).execute()
        except Exception:
            resp = q.execute()

        data = resp.data or []
        rows.extend(data)

        if len(data) < batch_size:
            break
        start += batch_size
    return rows


@st.cache_data(ttl=300)
def load_product_qty_sales_map() -> Tuple[List[str], Dict[str, Tuple[int, int]]]:
    products = _fetch_all_rows("products", "product_name,quantity")
    total_qty_map: Dict[str, int] = {}
    for r in products:
        name = r.get("product_name")
        if name:
            total_qty_map[name] = int(r.get("quantity") or 0)

    sales = _fetch_all_rows("daily_sales", "product_name,daily_sales_qty")
    sold_qty_map: Dict[str, int] = {}
    for r in sales:
        name = r.get("product_name")
        if name:
            sold_qty_map[name] = sold_qty_map.get(name, 0) + int(r.get("daily_sales_qty") or 0)

    names = sorted(set(total_qty_map) | set(sold_qty_map))
    stats: Dict[str, Tuple[int, int]] = {n: (total_qty_map.get(n, 0), sold_qty_map.get(n, 0)) for n in names}
    return names, stats


def format_product_option(option: str, stats: Dict[str, Tuple[int, int]]) -> str:
    if option in ("(선택 안 함)", "상품을 선택해주세요", "새로운 상품 입력") or not option:
        return option
    total, sold = stats.get(option, (0, 0))
    return f"{option} ({total:,}/{sold:,})"

def load_config_from_supabase():
    data = supabase.table("settings").select("*").execute().data
    cfg = {}
    for row in data:
        cfg[row["key"]] = float(row["value"])
    return cfg

config = load_config_from_supabase()

def can_save_daily_record(
    total_sales_qty: int,
    total_revenue: int,
    ad_sales_qty: int,
    ad_revenue: int,
    ad_cost: int,
) -> bool:
    # 판매(수량/매출) 또는 광고(광고비/전환수/광고매출) 중 하나라도 있으면 저장 허용
    has_sales = (int(total_sales_qty) > 0) or (int(total_revenue) > 0)
    has_ads = (int(ad_cost) > 0) or (int(ad_sales_qty) > 0) or (int(ad_revenue) > 0)
    return has_sales or has_ads

# 상품 정보 입력 상태 초기화 (탭2)
if "product_name_input" not in st.session_state: st.session_state["product_name_input_default"] = ""
if "sell_price_input" not in st.session_state: st.session_state.sell_price_input = ""
if "fee_rate_input" not in st.session_state: st.session_state.fee_rate_input = ""
if "inout_shipping_cost_input" not in st.session_state: st.session_state.inout_shipping_cost_input = ""
if "purchase_cost_input" not in st.session_state: st.session_state.purchase_cost_input = ""
if "quantity_input" not in st.session_state: st.session_state.quantity_input = ""
if "logistics_cost_input" not in st.session_state: st.session_state.logistics_cost_input = ""
if "customs_duty_input" not in st.session_state: st.session_state.customs_duty_input = ""
if "etc_cost_input" not in st.session_state: st.session_state.etc_cost_input = ""
if "is_edit_mode" not in st.session_state: st.session_state.is_edit_mode = False

# 일일 정산 입력 상태 초기화 (탭 2 number_input의 key를 사용)
if "total_sales_qty" not in st.session_state: st.session_state["total_sales_qty"] = 0
if "total_revenue" not in st.session_state: st.session_state["total_revenue"] = 0
if "ad_sales_qty" not in st.session_state: st.session_state["ad_sales_qty"] = 0
if "ad_revenue" not in st.session_state: st.session_state["ad_revenue"] = 0
if "ad_cost" not in st.session_state: st.session_state["ad_cost"] = 0
if "coupon_unit" not in st.session_state: st.session_state["coupon_unit"] = 0



def load_product_data(selected_product_name):
    if selected_product_name == "새로운 상품 입력":
        st.session_state.is_edit_mode = False
        st.session_state.product_name_input = ""
        st.session_state.sell_price_input = ""
        st.session_state.fee_rate_input = ""
        st.session_state.inout_shipping_cost_input = ""
        st.session_state.purchase_cost_input = ""
        st.session_state.quantity_input = ""
        st.session_state.logistics_cost_input = ""
        st.session_state.customs_duty_input = ""
        st.session_state.etc_cost_input = ""
    else:
        try:
            response = supabase.table("products").select("*").eq("product_name", selected_product_name).execute()
            if response.data:
                product_data = response.data[0]
                st.session_state.is_edit_mode = True

                st.session_state.product_name_input = product_data.get("product_name", "")

                def get_display_value(key, default=""):
                    val = product_data.get(key)
                    if val is None or val == 0:
                        return ""
                    if key == "fee":
                        return str(float(val))
                    return str(int(val)) if isinstance(val, (int, float)) and val == int(val) else str(val)

                st.session_state.sell_price_input = get_display_value("sell_price")
                st.session_state.fee_rate_input = get_display_value("fee")
                st.session_state.inout_shipping_cost_input = get_display_value("inout_shipping_cost")
                st.session_state.purchase_cost_input = get_display_value("purchase_cost")
                st.session_state.quantity_input = get_display_value("quantity")
                st.session_state.logistics_cost_input = get_display_value("logistics_cost")
                st.session_state.customs_duty_input = get_display_value("customs_duty")
                st.session_state.etc_cost_input = get_display_value("etc_cost")

        except Exception as e:
            st.error(f"상품 정보를 불러오는 중 오류가 발생했습니다: {e}")

def safe_int(value):
    try:
        if value in (None, ""):
            return 0
        return won(float(value))
    except (ValueError, TypeError):
        return 0

def safe_float(value):
    try:
        return float(value) if value else 0.0
    except (ValueError, TypeError):
        return 0.0

def won(x) -> int:
    """
    원 단위 금액 정수 확정(반올림).
    금액 라운딩은 전 탭에서 이것만 사용한다.
    """
    try:
        return int(round(float(x)))
    except (TypeError, ValueError):
        return 0

def cny_to_krw_float(cny_str: str, exchange_rate: float) -> float:
    """
    위안화 입력 → 원화 환산만 수행 (float 유지)
    중간 round / int / won 절대 금지
    """
    try:
        if cny_str is None:
            return 0.0
        s = str(cny_str).strip()
        if s == "":
            return 0.0
        return float(s) * float(exchange_rate)
    except (ValueError, TypeError):
        return 0.0

class ParsedCampaign:
    def __init__(self, campaign_name: str, status: str, ad_cost: int, ad_revenue: int, ad_sales_qty: int):
        self.campaign_name = campaign_name
        self.status = status
        self.ad_cost = ad_cost
        self.ad_revenue = ad_revenue
        self.ad_sales_qty = ad_sales_qty


def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

# ✅ 여기에 붙여넣기 (전역, 들여쓰기 0칸)
def _strip_edit_delete_suffix(text: str) -> str:
    s = _norm_ws(text)
    while True:
        before = s
        for suffix in ("수정", "삭제"):
            if s.endswith(suffix):
                s = s[: -len(suffix)].strip()
        if s == before:
            break
    return s.strip()

def _parse_won_like(text: str) -> int:
    t = _norm_ws(text)
    m = re.search(r"([\d,]+)", t)
    return int(m.group(1).replace(",", "")) if m else 0


def _parse_react_table(html_text: str):
    """
    bs4 없이 react-table(div.rt-*)에서 헤더/행을 추출한다.
    - headers: div.rt-th 텍스트
    - rows: div.rt-tr 내부의 div.rt-td[role="gridcell"] 텍스트
    """

    class _RTParser(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.headers = []
            self.rows = []

            self._stack = []  # markers: "row" | "hcell" | "dcell" | None
            self._in_cell = None  # "h" | "d" | None
            self._buf = []
            self._current_row = []

        @staticmethod
        def _get_attr(attrs, name):
            for k, v in attrs:
                if k == name:
                    return v
            return None

        def handle_starttag(self, tag, attrs):
            if tag != "div":
                self._stack.append(None)
                return

            cls = self._get_attr(attrs, "class") or ""
            role = self._get_attr(attrs, "role") or ""

            classes = set(cls.split())

            # row start
            if "rt-tr" in classes:
                self._stack.append("row")
                self._current_row = []
                return

            # header cell
            if "rt-th" in classes:
                self._stack.append("hcell")
                self._in_cell = "h"
                self._buf = []
                return

            # data cell
            if "rt-td" in classes and role == "gridcell":
                self._stack.append("dcell")
                self._in_cell = "d"
                self._buf = []
                return

            self._stack.append(None)

        def handle_data(self, data):
            if self._in_cell is not None:
                self._buf.append(data)

        def handle_endtag(self, tag):
            if not self._stack:
                return
            marker = self._stack.pop()

            if marker == "hcell":
                text = _norm_ws("".join(self._buf))
                self.headers.append(text)
                self._in_cell = None
                self._buf = []
                return

            if marker == "dcell":
                text = _norm_ws("".join(self._buf))
                self._current_row.append(text)
                self._in_cell = None
                self._buf = []
                return

            if marker == "row":
                if self._current_row:
                    self.rows.append(self._current_row)
                self._current_row = []
                return

    p = _RTParser()
    p.feed(html_text)
    p.close()

    headers = [h for h in p.headers if h]  # 빈 헤더 제거

    if not headers:
        raise ValueError("No headers found (rt-th)")

    # rows 길이 정렬
    rows = []
    for r in p.rows:
        if len(r) > len(headers):
            r = r[: len(headers)]
        elif len(r) < len(headers):
            r = r + [""] * (len(headers) - len(r))
        rows.append(r)

    return headers, rows

def parse_running_campaigns(html_text: str):
    headers, rows = _parse_react_table(html_text)

    def idx_of(pred):
        for i, h in enumerate(headers):
            if pred(h):
                return i
        raise KeyError("Required header not found")

    i_name = idx_of(lambda h: "캠페인 이름" in h or "상품명" in h)
    i_status = idx_of(lambda h: h == "상태" or "상태" in h)
    i_cost = idx_of(lambda h: "집행 광고비" in h)
    i_rev = idx_of(lambda h: "광고 전환 매출" in h)
    i_qty = idx_of(lambda h: "광고 전환 판매수" in h)

    out = []
    for r in rows:
        # 기존: status = r[i_status] (상태 열의 '운영 중' 텍스트 확인)
        # 변경: 행 전체 데이터(r) 중 'ON'이라는 텍스트가 포함된 요소가 있는지 확인
        if not any("ON" == str(cell).strip() for cell in r):
            continue

        name = _strip_edit_delete_suffix(r[i_name])
        status = r[i_status] # 상태 텍스트는 정보 저장용으로만 사용 

        if not name:
            continue

        out.append(
            ParsedCampaign(
                campaign_name=name,
                status=status,
                ad_cost=_parse_won_like(r[i_cost]),
                ad_revenue=_parse_won_like(r[i_rev]),
                ad_sales_qty=_parse_won_like(r[i_qty]),
            )
        )
    return out

def parse_sold_items_from_html(html_text: str) -> dict:
    import re
    from collections import defaultdict

    sold_pos = html_text.find('판매된 상품 목록')
    if sold_pos < 0:
        return {}

    section = html_text[sold_pos:sold_pos + 40000]
    tr_blocks = re.findall(r'<tr data-v-1c64ce3b="">(.*?)</tr>', section, re.DOTALL)

    agg = defaultdict(lambda: {'qty': 0, 'revenue': 0, 'options': 0})
    for tr in tr_blocks:
        name_m = re.search(r'<p data-v-1c64ce3b="">(.*?)</p>', tr)
        rev_m  = re.search(r'<td data-v-1c64ce3b="">([-\d,]+원)</td>', tr)
        qty_m  = re.search(r'<td data-v-1c64ce3b="">([-\d]+개)</td>', tr)
        if not name_m:
            continue
        name = name_m.group(1).strip()
        rev  = int(rev_m.group(1).replace(',', '').replace('원', '')) if rev_m else 0
        qty  = int(qty_m.group(1).replace('개', '')) if qty_m else 0
        bn   = name.split(',')[0].strip()
        agg[bn]['qty']     += qty
        agg[bn]['revenue'] += rev
        agg[bn]['options'] += 1

    return dict(agg)

def parse_sold_items_detail(html_text: str) -> list:
    import re
    sold_pos = html_text.find('판매된 상품 목록')
    if sold_pos < 0:
        return []
    section = html_text[sold_pos:sold_pos + 40000]
    tr_blocks = re.findall(r'<tr data-v-1c64ce3b="">(.*?)</tr>', section, re.DOTALL)
    results = []
    for tr in tr_blocks:
        name_m = re.search(r'<p data-v-1c64ce3b="">(.*?)</p>', tr)
        rev_m  = re.search(r'<td data-v-1c64ce3b="">([-\d,]+원)</td>', tr)
        qty_m  = re.search(r'<td data-v-1c64ce3b="">([-\d]+개)</td>', tr)
        if not name_m:
            continue
        full_name = name_m.group(1).strip()
        rev = int(rev_m.group(1).replace(',', '').replace('원', '')) if rev_m else 0
        qty = int(qty_m.group(1).replace('개', '')) if qty_m else 0
        results.append({
            'full_name': full_name,
            'base_name': full_name.split(',')[0].strip(),
            'qty': qty,
            'revenue': rev,
        })
    return results
    
def parse_product_ads(html_text: str):
    row_pattern = r'<div class="rt-tr[^"]*"[^>]*role="row">(.*?)(?=<div class="rt-tr-group|$)'
    row_htmls = re.findall(row_pattern, html_text, re.DOTALL)

    out = []
    for row_html in row_htmls[1:]:  # 0번은 헤더행
        on_off = re.findall(r'ant-switch-inner">(\w+)</span>', row_html)
        if not on_off or on_off[0] != "ON":
            continue
        names = re.findall(r'title="바로가기"[^>]*>([^<]+)</a>', row_html)
        name = names[0].strip() if names else ""
        if not name:
            continue

        # gridcell 값 순서대로 추출
        cells = re.findall(r'role="gridcell"[^>]*>.*?text--flex-ellipsis">(.*?)</div>', row_html, re.DOTALL)
        cells_clean = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]

        # [3]=광고전환판매수, [4]=광고전환매출, [6]=집행광고비
        qty  = _parse_won_like(cells_clean[3]) if len(cells_clean) > 3 else 0
        rev  = _parse_won_like(cells_clean[4]) if len(cells_clean) > 4 else 0
        cost = _parse_won_like(cells_clean[6]) if len(cells_clean) > 6 else 0

        out.append(ParsedCampaign(
            campaign_name=name,
            status="운영 중",
            ad_cost=cost,
            ad_revenue=rev,
            ad_sales_qty=qty,
        ))
    return out

def _yesterday_date() -> datetime.date:
    return datetime.date.today() - datetime.timedelta(days=1)


def _compute_daily(
    product_data: dict,
    report_date: datetime.date,
    product_name: str,
    total_sales_qty: int,
    total_revenue: int,
    coupon_unit: int,
    ad_sales_qty: int,
    ad_revenue_input: int,
    ad_cost: int,
):
    coupon_total = coupon_unit * total_sales_qty
    current_total_revenue = max(total_revenue - coupon_total, 0)

    ad_coupon_total = coupon_unit * ad_sales_qty
    ad_revenue_after_coupon = max(ad_revenue_input - ad_coupon_total, 0)

    organic_sales_qty = max(total_sales_qty - ad_sales_qty, 0)
    organic_revenue = max(current_total_revenue - ad_revenue_after_coupon, 0)

    quantity_val = product_data.get("quantity", 1) or 1
    quantity_for_calc = quantity_val if quantity_val > 0 else 1

    unit_purchase_cost = product_data.get("purchase_cost", 0) / quantity_for_calc
    unit_logistics = product_data.get("logistics_cost", 0) / quantity_for_calc
    unit_customs = product_data.get("customs_duty", 0) / quantity_for_calc
    unit_etc = product_data.get("etc_cost", 0) / quantity_for_calc
    fee_rate_db = product_data.get("fee", 0.0)

    daily_profit = (
        current_total_revenue
        - (current_total_revenue * fee_rate_db / 100 * 1.1)
        - (unit_purchase_cost * total_sales_qty)
        - (product_data.get("inout_shipping_cost", 0) * total_sales_qty * 1.1)
        - (unit_logistics * total_sales_qty)
        - (unit_customs * total_sales_qty)
        - (unit_etc * total_sales_qty)
        - (ad_cost * 1.1)
    )
    daily_profit = won(daily_profit)

    base_unit_cost = unit_purchase_cost + unit_logistics + unit_customs + unit_etc
    invest_for_day = base_unit_cost * total_sales_qty
    daily_roi = round(daily_profit / invest_for_day * 100, 2) if invest_for_day > 0 else 0

    data_to_save = {
        "date": report_date.isoformat(),
        "product_name": product_name,
        "daily_sales_qty": total_sales_qty,
        "daily_revenue": current_total_revenue,
        "ad_sales_qty": ad_sales_qty,
        "ad_revenue": ad_revenue_after_coupon,
        "organic_sales_qty": organic_sales_qty,
        "organic_revenue": organic_revenue,
        "daily_ad_cost": ad_cost,
        "daily_profit": daily_profit,
        "daily_roi": daily_roi,
    }

    return daily_profit, daily_roi, data_to_save

def cny_total_cost_krw(cny_str: str, exchange_rate: float, qty: int) -> int:
    """
    위안화 기준 총원가
    원 단위 정수 확정(won)은 여기서 단 1회만
    """
    return won(cny_to_krw_float(cny_str, exchange_rate) * int(qty))

def validate_inputs():
    required_fields = {
        "product_name_input": "상품명",
        "sell_price_input": "판매가",
        "fee_rate_input": "수수료율",
        "inout_shipping_cost_input": "입출고/배송비",
        "purchase_cost_input": "매입비",
        "quantity_input": "수량",
        "logistics_cost_input": "물류비",
        "customs_duty_input": "관세",
    }

    for key, name in required_fields.items():
        if not st.session_state.get(key):
            st.warning(f"**{name}** 필드를 채워주세요")
            return False

    return True

# --- [New Functions for ] ---
def calculate_profit_for_period(start_date: datetime.date, end_date: datetime.date, supabase: Client) -> int:
    """Supabase에서 지정된 기간 동안의 모든 상품의 총 순이익을 계산합니다."""
    start_str = start_date.isoformat()
    end_str = end_date.isoformat()
    
    try:
        # daily_sales 테이블에서 지정된 날짜 범위의 daily_profit만 가져옴
        response = supabase.table("daily_sales").select("daily_profit") \
            .gte("date", start_str) \
            .lte("date", end_str) \
            .execute()

        if response.data:
            df = pd.DataFrame(response.data)
            # daily_profit이 int/float형인지 확인하고 합산
            profit_sum = df["daily_profit"].sum() if "daily_profit" in df.columns else 0
            return int(profit_sum)
        return 0
    except Exception as e:
        # Supabase 연동 오류 발생 시 기본값 0 반환
        return 0

def get_date_range(period: str) -> tuple[datetime.date, datetime.date]:
    """오늘을 포함한 지정된 기간의 시작일과 종료일(오늘)을 반환합니다."""
    today = datetime.date.today()
    
    if period == "today": # 오늘
        return today, today
    elif period == "yesterday": # 어제
        yesterday = today - datetime.timedelta(days=1)
        return yesterday, yesterday
    elif period == "7days":
        # 오늘 포함 7일: 오늘 - 6일 = 시작일
        start_date = today - datetime.timedelta(days=6)
        return start_date, today
    elif period == "15days":
        # 오늘 포함 15일: 오늘 - 14일 = 시작일
        start_date = today - datetime.timedelta(days=14)
        return start_date, today
    elif period == "30days":
        # 오늘 포함 30일: 오늘 - 29일 = 시작일
        start_date = today - datetime.timedelta(days=29)
        return start_date, today
    elif period == "90days": # 90일 (기존 3months 대체)
        # 오늘 포함 90일: 오늘 - 89일 = 시작일
        start_date = today - datetime.timedelta(days=89) 
        return start_date, today
    elif period == "180days": # 180일
        start_date = today - datetime.timedelta(days=179)
        return start_date, today
    elif period == "365days": # 365일
        start_date = today - datetime.timedelta(days=364)
        return start_date, today
    else:
        return today, today # 기본값

# Note: display_profit_metric 함수는 박스형 출력 요청이 없어 제거되었습니다.
# --- [End of New Functions] ---


# =========================
# Sourcing Tool (Tab6)
# =========================

def _s_norm(s):
    return re.sub(r"\s+", " ", str(s or "").replace("\n", " ")).strip()


def _s_to_int(x, default=0):
    try:
        if x in (None, ""):
            return default
        return int(float(x))
    except (TypeError, ValueError):
        return default


def _s_to_float(x, default=0.0):
    try:
        if x in (None, ""):
            return default
        return float(x)
    except (TypeError, ValueError):
        return default


def _s_is_brand_x(v):
    return _s_norm(v).upper() in {"X", "엑스"}


def _s_is_shopping_o(v):
    return _s_norm(v).upper() in {"O", "0", "오"}


_S_MONTH_RE = re.compile(r"(?:^|[^0-9])((?:1[0-2])|(?:[1-9]))(?:[^0-9]|$)")


def _s_extract_months(text):
    """'3', '2,7,11', '5 6 7 8' 형태만 지원 ('3~5' 없음)."""
    t = _s_norm(text)
    months = set()
    if not t or ("없음" in t):
        return months
    for mm in _S_MONTH_RE.findall(t):
        try:
            m = int(mm)
            if 1 <= m <= 12:
                months.add(m)
        except ValueError:
            continue
    return months


@dataclass(frozen=True)
class SourcingCriteria:
    min_coupang_price: int = 10_000
    max_coupang_price: int = 30_000
    min_coupang_avg_reviews: float = 100.0
    max_coupang_avg_reviews: float = 500.0
    selected_months: frozenset[int] = frozenset()  # empty => 비시즌(계절성=없음)


def _s_pick_sheet(wb):
    return "all" if "all" in wb.sheetnames else wb.sheetnames[0]


def _s_build_cols(header_row):
    cols = {}
    for i, h in enumerate(header_row):
        name = _s_norm(h)
        if name and name not in cols:
            cols[name] = i
    return cols


def _s_find_first(cols, candidates):
    for c in candidates:
        if c in cols:
            return c
    return None


def _s_find_contains(cols, must_include, must_exclude=()):
    for name in cols.keys():
        if all(t in name for t in must_include) and all(t not in name for t in must_exclude):
            return name
    return None


def _s_find_tokens(cols, include_all=(), exclude_any=()):
    for name in cols.keys():
        if exclude_any and any(x in name for x in exclude_any):
            continue
        if include_all and not all(x in name for x in include_all):
            continue
        return name
    return None


def parse_sourcing_xlsx_stream(xlsx_path: str, criteria: SourcingCriteria, sheet_name: str | None = None):
    """
    XLSX 대용량 스트리밍 파서(openpyxl read_only).

    시즌 선택:
    - 월 선택 0개 => 비시즌(계절성 == '없음')
    - 월 선택 1개 이상 => 시즌(계절성 == '있음' AND 계절성 월에 선택월 포함)
    """
    if any((m < 1 or m > 12) for m in criteria.selected_months):
        raise ValueError(f"selected_months must be 1..12: {sorted(criteria.selected_months)}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = sheet_name or _s_pick_sheet(wb)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet}, available={wb.sheetnames}")
    ws = wb[sheet]

    it = ws.iter_rows(values_only=True)
    header = next(it)
    cols = _s_build_cols(header)

    col_keyword = _s_find_first(cols, ["키워드"])
    if not col_keyword:
        raise ValueError("Required column missing: 키워드")

    col_brand = _s_find_contains(cols, ["브랜드", "키워드"])
    col_shopping = _s_find_contains(cols, ["쇼핑성", "키워드"])

    col_cp_price = _s_find_tokens(cols, include_all=["쿠팡", "평균가"])
    col_cp_avg_reviews = _s_find_tokens(cols, include_all=["쿠팡", "평균리뷰수"])
    col_cp_total_reviews = _s_find_tokens(cols, include_all=["쿠팡", "총리뷰수"])
    col_cp_exposed_products = _s_find_tokens(cols, include_all=["쿠팡", "노출상품수"])

    col_last_year_total = _s_find_tokens(cols, include_all=["작년", "검색량"], exclude_any=["최대"])
    col_last_year_max_month = _s_find_tokens(cols, include_all=["작년", "최대", "검색", "월"], exclude_any=["검색량"])
    col_last_year_max_month_volume = _s_find_tokens(cols, include_all=["작년", "최대", "검색", "월", "검색량"])

    col_seasonality = _s_find_first(cols, ["계절성"])
    col_seasonality_month = _s_find_tokens(cols, include_all=["계절성", "월"])

    offseason_mode = not criteria.selected_months

    def get(row, col_name):
        if not col_name:
            return None
        idx = cols.get(col_name)
        return row[idx] if idx is not None and idx < len(row) else None

    def coupang_avg_reviews(row):
        v = get(row, col_cp_avg_reviews)
        if v not in (None, ""):
            return _s_to_float(v, 0.0)
        total = _s_to_float(get(row, col_cp_total_reviews), 0.0)
        denom = _s_to_float(get(row, col_cp_exposed_products), 0.0)
        return (total / denom) if denom > 0 else 0.0

    def seasonality_pass(row):
        s = _s_norm(get(row, col_seasonality))
        if not s:
            return False

        if "없음" in s:
            return offseason_mode

        if "있음" in s or s.upper() in {"O", "Y", "YES", "TRUE"}:
            if offseason_mode:
                return False
            months = _s_extract_months(get(row, col_seasonality_month))
            if not months:
                return False
            return bool(months.intersection(criteria.selected_months))

        return False

    def last_year_max_month(row):
        m = int(float(get(row, col_last_year_max_month) or 0))
        if not (1 <= m <= 12):
            raise ValueError(f"Invalid last-year max month: {m}")
        return m

    results = []
    seen = set()

    for row in it:
        keyword = _s_norm(get(row, col_keyword))
        if not keyword or keyword in seen:
            continue
        seen.add(keyword)

        if col_brand and not _s_is_brand_x(get(row, col_brand)):
            continue
        if col_shopping and not _s_is_shopping_o(get(row, col_shopping)):
            continue

        price = _s_to_int(get(row, col_cp_price), 0) if col_cp_price else 0
        if not (criteria.min_coupang_price <= price <= criteria.max_coupang_price):
            continue

        avg_rev = coupang_avg_reviews(row)
        if not (criteria.min_coupang_avg_reviews <= avg_rev <= criteria.max_coupang_avg_reviews):
            continue

        if not seasonality_pass(row):
            continue

        results.append(
            {
                "키워드": keyword,
                "작년 검색량": _s_to_int(get(row, col_last_year_total), 0),
                "작년 최대검색월": last_year_max_month(row),
                "작년최대 검색월 검색량": _s_to_int(get(row, col_last_year_max_month_volume), 0),
                "쿠팡 평균가": price,
                "쿠팡 평균 리뷰수": round(avg_rev, 2),
            }
        )

    return results


def _s_toggle_button(label: str, pressed: bool, key: str) -> bool:
    shown = f"✅ {label}" if pressed else label
    if st.button(shown, key=key):
        return not pressed
    return pressed


def _s_month_selector(key: str = "sourcing_months") -> set[int]:
    """
    12개 버튼(1~12월) 토글. 기본(empty)=비시즌.
    """
    if key not in st.session_state:
        st.session_state[key] = set()

    months = set(st.session_state[key])

    st.markdown("### 월 선택 (복수 선택)")
    cols = st.columns(6)
    for m in range(1, 13):
        col = cols[(m - 1) % 6]
        selected = m in months
        next_selected = _s_toggle_button(f"{m}월", selected, key=f"{key}_{m}")
        if next_selected != selected:
            if next_selected:
                months.add(m)
            else:
                months.remove(m)

    st.session_state[key] = months
    st.caption("모드: 시즌(있음)" if months else "모드: 비시즌(없음)")
    return months


def render_sourcing_tab():
    st.subheader("🧰 소싱툴 (엑셀 파싱)")

    months = _s_month_selector()
    uploaded = st.file_uploader("엑셀 업로드(.xlsx)", type=["xlsx"], key="sourcing_xlsx_uploader")

    if st.button("파싱 실행", type="primary", key="sourcing_run"):
        if uploaded is None:
            st.warning("파일 업로드가 필요합니다.")
            st.stop()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded.getbuffer())
            xlsx_path = tmp.name

        crit = SourcingCriteria(selected_months=frozenset(months))
        rows = parse_sourcing_xlsx_stream(xlsx_path, criteria=crit, sheet_name="all")

        st.success(f"완료: {len(rows)} rows")
        st.dataframe(rows, use_container_width=True)

        import pandas as pd
        df = pd.DataFrame(rows)
        st.download_button(
            "CSV 다운로드",
            data=df.to_csv(index=False).encode("utf-8-sig"),
            file_name="sourcing_filtered.csv",
            mime="text/csv",
            key="sourcing_download",
        )

def main():
    if 'show_product_info' not in st.session_state:
        st.session_state.show_product_info = False

    # 원본 파일의 코드를 4개의 탭으로 분리했습니다.
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["간단 마진계산기", "상품 정보 입력", "일일정산", "판매현황", "광고분석"])
    with tab1:  # 간단 마진 계산기 탭

        # 🔹 바깥 2컬럼: 왼쪽은 설정값 패널(가짜 사이드바), 오른쪽은 기존 계산 UI
        c1, c2, c3, c4, c5 = st.columns([0.5, 0.2, 1, 0.2, 1])

        # === 1) 탭1에서만 보이는 설정값 패널 ===
        with c1:
            st.markdown("### 🛠️ 설정값")

            config["FEE_RATE"]       = st.number_input("수수료율 (%)",       value=config.get("FEE_RATE", 10.8), step=0.1, format="%.2f")
            config["AD_RATE"]        = st.number_input("광고비율 (%)",       value=config.get("AD_RATE", 20.0),  step=0.1, format="%.2f")
            config["INOUT_COST"]     = st.number_input("입출고비용 (원)",    value=int(config.get("INOUT_COST", 3000)), step=100)
            config["PICKUP_COST"]    = st.number_input("회수비용 (원)",      value=int(config.get("PICKUP_COST", 0)),    step=100)
            config["RESTOCK_COST"]   = st.number_input("재입고비용 (원)",    value=int(config.get("RESTOCK_COST", 0)),   step=100)
            config["RETURN_RATE"]    = st.number_input("반품률 (%)",         value=config.get("RETURN_RATE", 0.0), step=0.1, format="%.2f")
            config["ETC_RATE"]       = st.number_input("기타비용률 (%)",     value=config.get("ETC_RATE", 2.0),  step=0.1, format="%.2f")
            config["EXCHANGE_RATE"]  = st.number_input("위안화 환율",        value=int(config.get("EXCHANGE_RATE", 300)), step=1)
            config["PACKAGING_COST"] = st.number_input("포장비 (원)",        value=int(config.get("PACKAGING_COST", 0)), step=100)
            config["GIFT_COST"]      = st.number_input("사은품 비용 (원)",   value=int(config.get("GIFT_COST", 0)),      step=100)

            if st.button("📂 기본값으로 저장", key="save_settings_tab1"):
                for k, v in config.items():
                    supabase.table("settings").upsert({"key": k, "value": v}).execute()
                st.success("Supabase에 저장 완료 ✅")

        # === 2) 오른쪽: 기존 탭1 UI (계산기) 그대로 ===
        with c3:
                st.markdown("<div style='margin-left:40px;'>", unsafe_allow_html=True)
                st.subheader("판매정보 입력")
                sell_price_raw = st.text_input("판매가 (원)", key="sell_price_raw")
                margin_display = st.empty()

                # 탭 1 마진 계산 로직
                if sell_price_raw.strip():
                    try:
                        target_margin = 50.0
                        sell_price_val = safe_int(sell_price_raw)
                        vat = 1.1
                        fee = won((sell_price_val * config['FEE_RATE'] / 100) * vat)
                        ad_fee = won((sell_price_val * config['AD_RATE'] / 100) * vat)
                        inout_cost = won(config['INOUT_COST'] * vat)
                        return_cost = won((config['PICKUP_COST'] + config['RESTOCK_COST']) * (config['RETURN_RATE'] / 100) * vat)
                        etc_cost = won((sell_price_val * config['ETC_RATE'] / 100) * vat)
                        packaging_cost = won(config['PACKAGING_COST'] * vat)
                        gift_cost = won(config['GIFT_COST'] * vat)
                        C_total_fixed_cost = fee + inout_cost + packaging_cost + gift_cost
                        raw_cost2 = sell_price_val * (1 - target_margin / 100) - C_total_fixed_cost
                        target_cost = max(0, won(raw_cost2))
                        yuan_cost = round((target_cost / config['EXCHANGE_RATE']) , 2)
                        profit = sell_price_val - (
                            target_cost + fee + inout_cost + packaging_cost + gift_cost
                        )
                        margin_display.markdown(
                            f"""
<div style='height:10px; line-height:10px; color:#f63366; font-size:15px; margin-bottom:15px;'>
    마진율 {int(target_margin)}% 기준: {format_number(target_cost)}원 ({yuan_cost:.2f}위안) / 마진: {format_number(profit)}원
</div>
""", unsafe_allow_html=True)
                    except:
                        margin_display.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
                else:
                    margin_display.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
                
                col1, col2 = st.columns(2)
                with col1:
                    st.text_input("위안화 (¥)", key="unit_yuan")
                with col2:
                    st.text_input("원화 (₩)", key="unit_won")
                st.text_input("수량", key="qty_raw", value=st.session_state.get("qty_raw", ""))

                calc_col, reset_col = st.columns(2)
                if calc_col.button("계산하기"):
                    st.session_state["show_result"] = True
                if "show_result" not in st.session_state:
                    st.session_state["show_result"] = False
                reset_col.button("리셋", on_click=reset_inputs)
                st.markdown("</div>", unsafe_allow_html=True)

            # --- 오른쪽: 결과 영역 ---
        with c5:
                # 탭 1 결과 출력 로직
                if st.session_state["show_result"]:
                    sell_price = safe_int(st.session_state.get("sell_price_raw", 0))
                    qty = safe_int(st.session_state.get("qty_raw", 1))

                    if sell_price <= 0 or qty <= 0:
                        st.warning("판매가와 수량을 정확히 입력해주세요.")
                        return
                    
                    # 원가 계산
                    unit_won_val = st.session_state.get("unit_won")
                    unit_yuan_val = st.session_state.get("unit_yuan")

                    # 단가(실수 유지: 중간 라운딩 금지)
                    if unit_won_val and unit_won_val.strip() != "":
                        unit_cost_val = float(unit_won_val)
                        cost_display = ""
                    elif unit_yuan_val and unit_yuan_val.strip() != "":
                        unit_cost_val = cny_to_krw_float(unit_yuan_val, config["EXCHANGE_RATE"])  # 환산만(float)
                        cost_display = f"{unit_yuan_val}위안"
                    else:
                        unit_cost_val = 0.0
                        cost_display = ""

                    # 총원가: 원화 확정은 한 곳에서 단 1회만
                    unit_cost = won(unit_cost_val * qty)  # 기존 라인 유지(결과 동일)
                    # 또는 아래로 완전 고정(추천: 실수로 다른 곳에서 won 하지 못하게)
                    # unit_cost = cny_total_cost_krw(unit_yuan_val, config["EXCHANGE_RATE"], qty) if (unit_yuan_val and unit_yuan_val.strip() != "" and not (unit_won_val and unit_won_val.strip() != "")) else won(unit_cost_val * qty)

                    
                    # 비용 계산
                    vat = 1.1
                    fee = won((sell_price * config["FEE_RATE"] / 100) * vat)
                    ad = won((sell_price * config["AD_RATE"] / 100) * vat)
                    inout = won(config["INOUT_COST"] * vat)
                    pickup = won(config["PICKUP_COST"])
                    restock = won(config["RESTOCK_COST"])
                    return_cost = won((pickup + restock) * (config["RETURN_RATE"] / 100) * vat)
                    etc = won((sell_price * config["ETC_RATE"] / 100) * vat)
                    packaging = won(config["PACKAGING_COST"] * vat)
                    gift = won(config["GIFT_COST"] * vat)
                    total_cost = unit_cost + fee + ad + inout + return_cost + etc + packaging + gift
                    profit2 = sell_price - total_cost  # 광고 포함 순이익(=최소 이익)
                    margin_profit = sell_price - (unit_cost + fee + inout + packaging + gift + etc)  # 광고 제외 마진
                    margin_ratio = round((margin_profit / sell_price) * 100, 2) if sell_price > 0 else 0
                    min_margin_ratio = round((profit2 / sell_price) * 100, 2) if sell_price > 0 else 0

                    # 투자금액(분모): 상품에 묶이는 돈만 (광고비 제외)
                    roi_invest_base = unit_cost + packaging + gift + etc + return_cost

                    # ROI(광고 없이): profit2는 '광고 포함'이므로 ad만 되돌려서 광고 제거
                    profit_no_ad = profit2 + ad
                    roi = round((profit_no_ad / roi_invest_base) * 100, 2) if roi_invest_base > 0 else 0

                    # 최소 ROI(광고 포함)
                    min_roi = round((profit2 / roi_invest_base) * 100, 2) if roi_invest_base > 0 else 0

                    # 손익분기 ROAS: "정밀(exact)" 기준으로 통일 (혼합 제거)
                    unit_cost_exact = unit_cost_val * qty
                    fee_exact = won(sell_price * config["FEE_RATE"] / 100 * vat)
                    inout_exact = won(config["INOUT_COST"] * vat)
                    etc_exact = won(sell_price * config["ETC_RATE"] / 100 * vat)
                    packaging_exact = won(config["PACKAGING_COST"] * vat)
                    gift_exact = won(config["GIFT_COST"] * vat)


                    margin_profit_exact = sell_price - (
                        unit_cost_exact
                        + fee_exact
                        + inout_exact
                        + packaging_exact
                        + gift_exact
                        + etc_exact
                    )

                    be_roas = round((sell_price / margin_profit_exact) * 100, 2) if margin_profit_exact > 0 else 0


                    col_title, col_button = st.columns([4,1])
                    with col_title:
                        st.markdown("### 📊 계산 결과")

                    if cost_display:
                        st.markdown(f"- 🏷️ **원가:** {format_number(unit_cost)}원 ({cost_display})" if unit_cost > 0 else f"- 🏷️ **원가:** {format_number(unit_cost)}원")
                    else:
                        st.markdown(f"- 🏷️ **원가:** {format_number(unit_cost)}원")
                    st.markdown(f"- 💰 **마진:** {format_number(margin_profit)}원 / ROI: {roi:.2f}%")
                    st.markdown(f"- 📈 **마진율:** {margin_ratio:.2f}%")
                    st.markdown(f"- 🧾 **최소 이익:** {format_number(profit2)}원 / ROI: {min_roi:.2f}%")
                    st.markdown(f"- 📉 **최소마진율:** {(profit2 / sell_price * 100):.2f}%")
                    st.markdown(f"- 📊 **손익분기 ROAS:** {be_roas:.2f}%")

                    with st.expander("📦 상세 비용 항목 보기", expanded=False):
                        def styled_line(label, value):
                            return f"<div style='font-size:15px;'><strong>{label}</strong> {value}</div>"
                        st.markdown(styled_line("판매가:", f"{format_number(sell_price)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("원가:", f"{format_number(unit_cost)}원 ({cost_display})" if cost_display else f"{format_number(unit_cost)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("수수료:", f"{format_number(fee)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("광고비:", f"{format_number(ad)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("입출고비용:", f"{format_number(inout)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("회수비용:", f"{format_number(pickup)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("재입고비용:", f"{format_number(restock)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("반품비용:", f"{format_number(return_cost)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("기타비용:", f"{format_number(etc)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("포장비:", f"{format_number(packaging)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("사은품 비용:", f"{format_number(gift)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("총비용:", f"{format_number(total_cost)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("최소 이익:", f"{format_number(profit2)}원"), unsafe_allow_html=True)
                        st.markdown(styled_line("최소마진율:", f"{(profit2/sell_price*100):.2f}%"), unsafe_allow_html=True)
                        st.markdown(styled_line("투자수익률:", f"{roi:.2f}%"), unsafe_allow_html=True)


    with tab2: # 원본 파일의 '세부 마진 계산기' 탭 내부의 '상품 정보 입력' 내용
        c1, c2, c3 = st.columns([1, 1, 1])     
        with c2:
                st.subheader("상품 정보 입력")
        
                # 상품 목록 로드
                product_list = ["새로운 상품 입력"]
                try:
                    response = supabase.table("products").select("product_name").order("product_name").execute()
                    if response.data:
                        saved_products = [item['product_name'] for item in response.data]
                        product_list.extend(saved_products)
                except Exception as e:
                    st.error(f"상품 목록을 불러오는 중 오류가 발생했습니다: {e}")

                st.selectbox(
                    "저장된 상품 선택 또는 새로 입력",
                    product_list,
                    key="product_loader",
                    on_change=lambda: load_product_data(st.session_state.product_loader),
                )

                st.text_input(
                    "상품명",
                    value=st.session_state.get("product_name_input_default", ""),
                    key="product_name_input",
                    placeholder="예: 무선 이어폰"
                )


                # 상품 세부 정보 입력
                col_left, col_right = st.columns(2)
                with col_left:
                    st.text_input("판매가", key="sell_price_input")
                with col_right:
                    st.text_input("수수료율 (%)", key="fee_rate_input")
                with col_left:
                    st.text_input("입출고/배송비", key="inout_shipping_cost_input")
                with col_right:
                    st.text_input("매입비", key="purchase_cost_input")
                with col_left:
                    st.text_input("수량", key="quantity_input")

                sell_price = safe_int(st.session_state.sell_price_input)
                fee_rate = safe_float(st.session_state.fee_rate_input)
                inout_shipping_cost = safe_int(st.session_state.inout_shipping_cost_input)
                purchase_cost = safe_int(st.session_state.purchase_cost_input)
                quantity = safe_int(st.session_state.quantity_input)

                quantity_for_calc = quantity if quantity > 0 else 1

                with col_right:
                    try:
                        unit_purchase_cost = purchase_cost / quantity_for_calc
                    except (ZeroDivisionError, TypeError):
                        unit_purchase_cost = 0
                    st.text_input("매입단가", value=f"{unit_purchase_cost:,.0f}원", disabled=True)
                with col_left:
                    st.text_input("물류비", key="logistics_cost_input")
                with col_right:
                    st.text_input("관세", key="customs_duty_input")

                st.text_input("기타", key="etc_cost_input")

# --- 실시간 수익성 분석 (소수점 절사 및 정수 출력 버전) ---
                st.markdown("---")
                st.subheader("📊 실시간 수익성 분석 (예측)")
                
                try:
                    def get_clean_val(key):
                        val = st.session_state.get(key, "0")
                        if isinstance(val, str):
                            val = val.replace(",", "").replace("원", "").strip()
                        try:
                            return float(val) if val else 0.0
                        except:
                            return 0.0

                    s_p = get_clean_val("sell_price_input")      # 판매가
                    f_r = get_clean_val("fee_rate_input")       # 수수료율
                    i_c = get_clean_val("inout_shipping_cost_input") # 입출고비
                    p_c = get_clean_val("purchase_cost_input")   # 총매입비
                    qty_v = get_clean_val("quantity_input")     # 수량
                    l_c = get_clean_val("logistics_cost_input")  # 총물류비
                    c_d = get_clean_val("customs_duty_input")    # 총관세
                    e_c = get_clean_val("etc_cost_input")        # 총기타비용

                    vat_v = 1.1
                    q_calc = qty_v if qty_v > 0 else 1
                    
                    # 단위당 원가 (소수점 이하 절사)
                    u_p = int(p_c / q_calc)
                    u_l = int(l_c / q_calc)
                    u_c = int(c_d / q_calc)
                    u_e = int(e_c / q_calc)
                    unit_invest = u_p + u_l + u_c + u_e

                    # 수수료 및 배송비 (VAT 1.1 적용 후 소수점 이하 절사)
                    actual_fee = int(s_p * (f_r / 100) * vat_v)
                    actual_inout = int(i_c * vat_v)
                    
                    # 마진 계산
                    margin_p = int(s_p - actual_fee - actual_inout - unit_invest)
                    
                    # 지표 산출 (모두 정수형으로 변환)
                    # 마진율
                    m_ratio = int((margin_p / s_p * 100)) if s_p > 0 else 0
                    # ROI
                    roi_v = int((margin_p / unit_invest * 100)) if unit_invest > 0 else 0
                    # 손익분기 ROAS
                    be_roas_v = int((s_p / margin_p * 100)) if margin_p > 0 else 0

                    # 화면 출력 (소수점 없이 정수+% 로만 표시)
                    m_col1, m_col2, m_col3 = st.columns(3)
                    m_col1.metric("마진율", f"{m_ratio}%")
                    m_col2.metric("ROI", f"{roi_v}%")
                    m_col3.metric("손익분기 ROAS", f"{be_roas_v}%")
                                        
                except Exception:
                    pass
                
                st.markdown("---")
            
                logistics_cost = safe_int(st.session_state.logistics_cost_input)
                customs_duty = safe_int(st.session_state.customs_duty_input)
                etc_cost = safe_int(st.session_state.etc_cost_input)

                quantity_to_save = quantity
        
                # 저장/수정/삭제 버튼 로직
                if st.session_state.is_edit_mode:
                    col_mod, col_del = st.columns(2)

                    with col_mod:
                        if st.button("수정하기"):
                            if validate_inputs():
                                try:
                                    old_name = st.session_state.product_loader
                                    new_name = st.session_state.product_name_input

                                    data_to_update = {
                                        "product_name": new_name,
                                        "sell_price": safe_int(st.session_state.sell_price_input),
                                        "fee": safe_float(st.session_state.fee_rate_input),
                                        "inout_shipping_cost": safe_int(st.session_state.inout_shipping_cost_input),
                                        "purchase_cost": safe_int(st.session_state.purchase_cost_input),
                                        "quantity": safe_int(st.session_state.quantity_input),
                                        "unit_purchase_cost": (
                                            safe_int(st.session_state.purchase_cost_input) / max(safe_int(st.session_state.quantity_input), 1)
                                        ),
                                        "logistics_cost": safe_int(st.session_state.logistics_cost_input),
                                        "customs_duty": safe_int(st.session_state.customs_duty_input),
                                        "etc_cost": safe_int(st.session_state.etc_cost_input),
                                    }

                                    if old_name != new_name:
                                        # ✅ 이름이 바뀐 경우: 기존 행 update
                                        supabase.rpc(
                                            "update_product_by_old_name",
                                            {"old_name": old_name, "p_data": data_to_update}
                                        ).execute()

                                        # ✅ daily_sales 테이블도 이름 동기화
                                        supabase.rpc(
                                            "update_daily_sales_name",
                                            {"old_name": old_name, "new_name": new_name}
                                        ).execute()
                                    else:
                                        # ✅ 이름이 같으면 기존 upsert 그대로
                                        supabase.rpc("upsert_product", {"p_data": data_to_update}).execute()
        
                                    st.success("데이터가 수정되었습니다!")
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"상품명 수정 중 오류가 발생했습니다: {e}")

                    with col_del:
                        if st.button("삭제하기"):
                            try:
                                product_to_delete = st.session_state.product_name_input
                                supabase.rpc("delete_product_and_sales", {"p_name": product_to_delete}).execute()
                                st.success(f"'{product_to_delete}' 상품과 관련된 모든 데이터가 삭제되었습니다!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"데이터 삭제 중 오류가 발생했습니다: {e}")

                else:            
                    if st.button("상품 저장하기"):
                        if validate_inputs():
                            try:
                                data_to_save = {
                                    "created_at": datetime.datetime.now().isoformat(),
                                    "product_name": st.session_state.product_name_input,
                                    "sell_price": safe_int(st.session_state.sell_price_input),
                                    "fee": safe_float(st.session_state.fee_rate_input),
                                    "inout_shipping_cost": safe_int(st.session_state.inout_shipping_cost_input),
                                    "purchase_cost": safe_int(st.session_state.purchase_cost_input),
                                    "quantity": safe_int(st.session_state.quantity_input),
                                    "unit_purchase_cost": (
                                        safe_int(st.session_state.purchase_cost_input) / max(safe_int(st.session_state.quantity_input), 1)
                                    ),
                                    "logistics_cost": safe_int(st.session_state.logistics_cost_input),
                                    "customs_duty": safe_int(st.session_state.customs_duty_input),
                                    "etc_cost": safe_int(st.session_state.etc_cost_input),
                                }
                                supabase.rpc("upsert_product", {"p_data": data_to_save}).execute()
                                st.success(f"'{st.session_state.product_name_input}' 상품이 저장(또는 수정)되었습니다!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"데이터 저장 중 오류가 발생했습니다: {e}")

    with tab3:  # 일일정산
        st.markdown("""
            <style>
            [data-testid="column"]:nth-of-type(2) {
                height: 80vh;
                overflow-y: auto;
            }
            [data-testid="column"]:nth-of-type(3) {
                position: sticky;
                top: 60px;
                height: fit-content;
            }
            </style>
        """, unsafe_allow_html=True)
        
        c1, c2, c3 = st.columns([1, 1, 1])
        with c2:
            st.subheader("일일 정산")

            uploaded_files = st.file_uploader(
                "쿠팡 HTML 업로드 (복수 선택 가능)",
                type=["html", "htm"],
                key="daily_html",
                accept_multiple_files=True
            )

            parsed_campaigns = []
            if uploaded_files:
                for uploaded_html in uploaded_files:
                    html_text = uploaded_html.getvalue().decode("utf-8", errors="ignore")
                    try:
                        is_type2 = not any("캠페인 이름" in h for h in _parse_react_table(html_text)[0])
                        if is_type2:
                            result = parse_product_ads(html_text)
                            st.success(f"[{uploaded_html.name}] 운영 중 상품 {len(result)}개 파싱 완료")
                        else:
                            result = parse_running_campaigns(html_text)
                            st.success(f"[{uploaded_html.name}] 운영 중 캠페인 {len(result)}개 파싱 완료")
                        parsed_campaigns.extend(result)
                    except Exception:
                        pass

            st.markdown("---")

            if uploaded_files:
                if not parsed_campaigns:
                    st.info("HTML 업로드하면 자동 입력됩니다. 업로드가 없거나 캠페인 0개면 아래 수동으로 입력하세요.")
                else:
                    # ... (생략) parsed_campaigns 생성 이후 ...
                    # (공통 날짜) 기본값: 어제
                    if "auto_report_date" not in st.session_state:
                        st.session_state["auto_report_date"] = _yesterday_date()

                    st.date_input("날짜 선택", key="auto_report_date")

                    # ✅ 상품명 + (전체/판매) 라벨용 데이터 1회 로드 (for-loop 밖)
                    try:
                        product_names, product_stats = load_product_qty_sales_map()
                    except Exception as e:
                        st.error(f"상품/판매수량 집계 로드 실패: {e}")
                        product_names, product_stats = [], {}

                    PRODUCT_PICKER_OPTIONS = ["(선택 안 함)"] + product_names

                    # ✅ 업로드 파일 단위 '제외' 상태 유지 (자동 입력 폼 전용)
                    upload_sig = ":".join([f"{f.name}:{f.size}" for f in uploaded_files])
                    excluded_state_key = f"auto_excluded_campaigns::{upload_sig}"
                    if excluded_state_key not in st.session_state:
                        st.session_state[excluded_state_key] = set()

                    # 상품 합산 사전 계산
                    sold_summary = {}
                    if uploaded_files:
                        for uf in uploaded_files:
                            ht = uf.getvalue().decode("utf-8", errors="ignore")
                            if '판매된 상품 목록' not in ht:
                                continue  # 광고 HTML은 건너뜀
                            parsed_sold = parse_sold_items_from_html(ht)
                            for bn, v in parsed_sold.items():
                                if bn not in sold_summary:
                                    sold_summary[bn] = {'qty': 0, 'revenue': 0, 'options': 0}
                                sold_summary[bn]['qty']     += v['qty']
                                sold_summary[bn]['revenue'] += v['revenue']
                                sold_summary[bn]['options'] += v['options']

                    # 정렬된 전체 아이템 (필터링 전)
                    product_options = ["전체"] + sorted(sold_summary.keys())
                    st.session_state.setdefault("sold_product_filter_global", "전체")
                    sorted_items_all = sorted(sold_summary.items(), key=lambda x: -x[1]['revenue'])       

                    for i, camp in enumerate(parsed_campaigns, start=1):
                        prefix = f"auto_{i}"

                        camp_key = f"{upload_sig}:{i}:{camp.campaign_name}"
                        is_excluded = camp_key in st.session_state[excluded_state_key]

                        if f"{prefix}_product_picker" not in st.session_state:
                            st.session_state[f"{prefix}_product_picker"] = "(선택 안 함)"

                        st.session_state.setdefault(f"{prefix}_total_sales_qty", 0)
                        st.session_state.setdefault(f"{prefix}_total_revenue", 0)
                        st.session_state.setdefault(f"{prefix}_coupon_unit", 0)

                        sig = (upload_sig, i, camp.campaign_name)

                        cur_qty = st.session_state.get(f"{prefix}_ad_sales_qty")
                        cur_rev = st.session_state.get(f"{prefix}_ad_revenue")
                        cur_cost = st.session_state.get(f"{prefix}_ad_cost")

                        need_refill = (
                            st.session_state.get(f"{prefix}_autofill_sig") != sig
                            or cur_qty in (None, 0)
                            or cur_rev in (None, 0)
                            or cur_cost in (None, 0)
                        )

                        if need_refill:
                            st.session_state[f"{prefix}_ad_sales_qty"] = int(camp.ad_sales_qty or 0)
                            st.session_state[f"{prefix}_ad_revenue"] = int(camp.ad_revenue or 0)
                            st.session_state[f"{prefix}_ad_cost"] = int(camp.ad_cost or 0)
                            st.session_state[f"{prefix}_autofill_sig"] = sig
                    # 상품 선택 + 합산 박스 (캠페인마다 반복)
                        if "sold_product_filter_global" not in st.session_state:
                            st.session_state["sold_product_filter_global"] = "전체"
                        
                        # global 값으로 먼저 덮어쓰기
                        if f"sold_product_filter_{i}" not in st.session_state:
                            st.session_state[f"sold_product_filter_{i}"] = st.session_state["sold_product_filter_global"]
                        
                        def update_global():
                            st.session_state["sold_product_filter_global"] = st.session_state["controller"]
                            st.rerun()

                        st.selectbox(
                            "📦 상품 선택",
                            product_options,
                            index=product_options.index(st.session_state["sold_product_filter_global"]),
                            key=f"display_{i}",
                            disabled=True,
                        )

                        # 값 동기화
                        if selected_product == "전체":
                            filtered_items = sorted_items_all
                        else:
                            filtered_items = [(bn, v) for bn, v in sold_summary.items() if bn == selected_product]
                        sorted_items = sorted(filtered_items, key=lambda x: -x[1]['revenue'])
                        
                        # 상품 선택 + 합산 박스 (캠페인마다 반복)
                        if sorted_items:
                            total_revenue = sum(v['revenue'] for _, v in sorted_items)
                            total_qty = sum(v['qty'] for _, v in sorted_items)

                            if selected_product != "전체":
                                detail_rows = []
                                for uf in uploaded_files:
                                    ht = uf.getvalue().decode("utf-8", errors="ignore")
                                    if '판매된 상품 목록' not in ht:
                                        continue
                                    for item in parse_sold_items_detail(ht):
                                        if item['base_name'] == selected_product:
                                            detail_rows.append(item)
                                detail_rows.sort(key=lambda x: -x['revenue'])
                                
                                rows = "".join([
                                    f"<div style='display:flex;justify-content:space-between;padding:2px 0;font-size:13px;'>"
                                    f"<span style='color:#666;'>{item['full_name'].split(',')[1].strip() if ',' in item['full_name'] else item['full_name']}</span>"
                                    f"<span><b>{item['qty']:,}개</b> | <b>{item['revenue']:,}원</b></span>"
                                    f"</div>"
                                    for item in detail_rows
                                ])
                            else:
                                rows = "".join([
                                    f"<div style='display:flex;justify-content:space-between;padding:2px 0;font-size:13px;'>"
                                    f"<span>{bn}</span>"
                                    f"<span><b>{v['qty']:,}개</b> | <b>{v['revenue']:,}원</b></span>"
                                    f"</div>"
                                    for bn, v in sorted_items
                                ])

                            html_block = (
                                "<div style='border:1px solid #dee2e6;border-radius:6px;"
                                "padding:8px 14px;margin-bottom:8px;background:#f8f9fa;font-size:13px;'>"
                                f"<div style='font-weight:bold;margin-bottom:6px;color:#555;'>💰 총 {total_revenue:,}원 &nbsp;|&nbsp; {total_qty:,}개</div>"
                                + rows +
                                "</div>"
                            )
                            st.markdown(html_block, unsafe_allow_html=True)
                            
                        with st.container(border=True):
                            
                            left, right = st.columns([8, 2])
                            with left:
                                st.markdown(f"#### {i}. {camp.campaign_name}")
                                st.caption(f"📅 적용 날짜: {st.session_state['auto_report_date']}")
                            with right:
                                if not is_excluded:
                                    if st.button("🗑️ 제외", key=f"{prefix}_exclude_btn"):
                                        st.session_state[excluded_state_key].add(camp_key)
                                        st.rerun()

                            if is_excluded:
                                st.caption("🚫 제외됨 (저장 제외)")
                                continue

                            st.selectbox(
                                "",
                                PRODUCT_PICKER_OPTIONS,
                                key=f"{prefix}_product_picker",
                                format_func=lambda x: format_product_option(x, product_stats),
                                label_visibility="collapsed",
                            )


                            st.markdown("#### 전체 판매")

                            st.number_input(
                                "전체 판매 수량",
                                min_value=0,
                                step=1,
                                format="%d",
                                key=f"{prefix}_total_sales_qty",
                            )
                            st.number_input(
                                "전체 매출액",
                                min_value=0,
                                step=1000,
                                format="%d",
                                key=f"{prefix}_total_revenue",
                            )
                            st.number_input(
                                "개당 쿠폰가 (원)",
                                min_value=0,
                                step=100,
                                format="%d",
                                key=f"{prefix}_coupon_unit",
                            )

                            st.markdown("#### 광고 판매 (HTML 자동채움)")
                            st.number_input(
                                "광고 전환 판매 수량",
                                min_value=0,
                                step=1,
                                format="%d",
                                key=f"{prefix}_ad_sales_qty",
                            )
                            st.number_input(
                                "광고 매출액",
                                min_value=0,
                                step=1000,
                                format="%d",
                                key=f"{prefix}_ad_revenue",
                            )
                            st.number_input(
                                "광고비용",
                                min_value=0,
                                step=1000,
                                format="%d",
                                key=f"{prefix}_ad_cost",
                            )

                            st.markdown("#### 자연 판매 (자동 계산)")
                            total_sales_qty = int(st.session_state.get(f"{prefix}_total_sales_qty", 0))
                            display_revenue = int(st.session_state.get(f"{prefix}_total_revenue", 0))
                            ad_sales_qty = int(st.session_state.get(f"{prefix}_ad_sales_qty", 0))
                            ad_revenue_input = int(st.session_state.get(f"{prefix}_ad_revenue", 0))
                            coupon_unit = int(st.session_state.get(f"{prefix}_coupon_unit", 0))

                            coupon_total = coupon_unit * total_sales_qty
                            actual_revenue = max(display_revenue - coupon_total, 0)

                            ad_coupon_total = coupon_unit * ad_sales_qty
                            ad_revenue_after_coupon = max(ad_revenue_input - ad_coupon_total, 0)

                            organic_sales_qty_calc = int(max(total_sales_qty - ad_sales_qty, 0))
                            organic_revenue_calc = int(max(actual_revenue - ad_revenue_after_coupon, 0))

                            st.session_state[f"{prefix}_organic_qty_view"] = organic_sales_qty_calc
                            st.session_state[f"{prefix}_organic_rev_view"] = organic_revenue_calc

                            st.number_input(
                                "자연 판매 수량",
                                min_value=0,
                                step=1,
                                format="%d",
                                disabled=True,
                                key=f"{prefix}_organic_qty_view",
                            )
                            st.number_input(
                                "자연 판매 매출액",
                                min_value=0,
                                step=1000,
                                format="%d",
                                disabled=True,
                                key=f"{prefix}_organic_rev_view",
                            )

                            # ✅ 일일 순이익 계산도 공통 날짜 사용
                            picked = (st.session_state.get(f"{prefix}_product_picker") or "").strip()
                            product_name = "" if picked in ("", "(선택 안 함)") else picked

                            total_sales_qty = int(st.session_state.get(f"{prefix}_total_sales_qty", 0))
                            total_revenue = int(st.session_state.get(f"{prefix}_total_revenue", 0))
                            coupon_unit = int(st.session_state.get(f"{prefix}_coupon_unit", 0))
                            ad_sales_qty = int(st.session_state.get(f"{prefix}_ad_sales_qty", 0))
                            ad_revenue_input = int(st.session_state.get(f"{prefix}_ad_revenue", 0))
                            ad_cost = int(st.session_state.get(f"{prefix}_ad_cost", 0))

                            if product_name and can_save_daily_record(total_sales_qty, total_revenue, ad_sales_qty, ad_revenue_input, ad_cost):
                                resp_prod = (
                                    supabase.table("products")
                                    .select("*")
                                    .eq("product_name", product_name)
                                    .execute()
                                )
                                if resp_prod.data:
                                    product_data = resp_prod.data[0]
                                    daily_profit, _, _ = _compute_daily(
                                        product_data=product_data,
                                        report_date=st.session_state["auto_report_date"],
                                        product_name=product_name,
                                        total_sales_qty=total_sales_qty,
                                        total_revenue=total_revenue,
                                        coupon_unit=coupon_unit,
                                        ad_sales_qty=ad_sales_qty,
                                        ad_revenue_input=ad_revenue_input,
                                        ad_cost=ad_cost,
                                    )
                                    st.metric(label="일일 순이익금", value=f"{daily_profit:,}원")

                            st.markdown("---")

                    # ✅ 전체 저장도 공통 날짜 사용
                    if st.button("전체 저장 (N건 일괄)", key="auto_save_all"):
                        errors = []
                        payloads = []
                        excluded = set(st.session_state.get(excluded_state_key, set()))
                        skipped = 0

                        for i, camp in enumerate(parsed_campaigns, start=1):
                            prefix = f"auto_{i}"
                            camp_key = f"{upload_sig}:{i}:{camp.campaign_name}"

                            if camp_key in excluded:
                                skipped += 1
                                continue

                            picked = (st.session_state.get(f"{prefix}_product_picker") or "").strip()
                            product_name = "" if picked in ("", "(선택 안 함)") else picked

                            report_date = st.session_state["auto_report_date"]  # ✅ 공통 날짜
                            total_sales_qty = int(st.session_state.get(f"{prefix}_total_sales_qty", 0))
                            total_revenue = int(st.session_state.get(f"{prefix}_total_revenue", 0))
                            coupon_unit = int(st.session_state.get(f"{prefix}_coupon_unit", 0))
                            ad_sales_qty = int(st.session_state.get(f"{prefix}_ad_sales_qty", 0))
                            ad_revenue_input = int(st.session_state.get(f"{prefix}_ad_revenue", 0))
                            ad_cost = int(st.session_state.get(f"{prefix}_ad_cost", 0))

                            if not product_name:
                                errors.append(f"[{i}] 상품명을 선택해주세요")
                                continue

                            if not can_save_daily_record(total_sales_qty, total_revenue, ad_sales_qty, ad_revenue_input, ad_cost):
                                errors.append(f"[{i}] 판매(수량/매출) 또는 광고(광고비/전환수/광고매출) 중 1개는 필요")
                                continue

                            response = (
                                supabase.table("products")
                                .select("*")
                                .eq("product_name", product_name)
                                .execute()
                            )
                            if not response.data:
                                errors.append(
                                    f"[{i}] products에 '{product_name}' 없음 (상품 정보 입력 탭에서 먼저 저장)"
                                )
                                continue

                            product_data = response.data[0]
                            daily_profit, daily_roi, data_to_save = _compute_daily(
                                product_data=product_data,
                                report_date=report_date,  # ✅ 공통 날짜
                                product_name=product_name,
                                total_sales_qty=total_sales_qty,
                                total_revenue=total_revenue,
                                coupon_unit=coupon_unit,
                                ad_sales_qty=ad_sales_qty,
                                ad_revenue_input=ad_revenue_input,
                                ad_cost=ad_cost,
                            )
                            payloads.append((i, report_date, daily_profit, daily_roi, data_to_save))

                        if errors:
                            st.error("저장 실패: 아래 항목 확인")
                            for e in errors:
                                st.write(f"- {e}")
                        else:
                            try:
                                for _, _, _, _, data_to_save in payloads:
                                    supabase.rpc("upsert_daily_sales", {"p_data": data_to_save}).execute()
                                st.success(f"{len(payloads)}건 저장 완료 ✅")
                                if skipped:
                                    st.info(f"제외 처리: {skipped}건")
                            except Exception as e:
                                st.error(f"저장 중 오류: {e}")

            else:
                # -------------------------
                # 수동 모드: 원본 tab3 로직 그대로 (변경 없음)
                # -------------------------
                try:
                    product_names, product_stats = load_product_qty_sales_map()
                except Exception as e:
                    st.error(f"상품/판매수량 집계 로드 실패: {e}")
                    product_names, product_stats = [], {}

                product_list = ["상품을 선택해주세요"] + product_names

                selected_product_name = st.selectbox(
                    "상품 선택",
                    product_list,
                    key="product_select_daily",
                    format_func=lambda x: format_product_option(x, product_stats),
                )

                product_data = {}
                if selected_product_name and selected_product_name != "상품을 선택해주세요":
                    try:
                        response = supabase.table("products").select("*").eq("product_name", selected_product_name).execute()
                        if response.data:
                            product_data = response.data[0]
                    except Exception as e:
                        st.error(f"상품 정보를 불러오는 중 오류가 발생했습니다: {e}")

                with st.expander("상품 상세 정보"):
                    if selected_product_name == "상품을 선택해주세요":
                        st.info("먼저 상품을 선택해주세요.")
                    elif product_data:
                        display_qty = product_data.get('quantity') or 0
                        qty = display_qty if display_qty > 0 else 1

                        sell_price = product_data.get("sell_price", 0) or 0
                        fee_rate = product_data.get("fee", 0.0) or 0.0
                        inout_shipping = product_data.get("inout_shipping_cost", 0) or 0

                        unit_purchase  = won(product_data.get("unit_purchase_cost", 0) or 0)
                        unit_logistics = won((product_data.get("logistics_cost", 0) or 0) / qty)
                        unit_customs   = won((product_data.get("customs_duty", 0) or 0) / qty)
                        unit_etc       = won((product_data.get("etc_cost", 0) or 0) / qty)

                        fee_per_unit   = won(sell_price * (fee_rate / 100) * 1.1)
                        inout_per_unit = won(inout_shipping * 1.1)

                        margin_profit_unit = sell_price - (
                            fee_per_unit
                            + inout_per_unit
                            + unit_purchase
                            + unit_logistics
                            + unit_customs
                            + unit_etc
                        )

                        margin_rate_pct = (margin_profit_unit / sell_price * 100) if sell_price > 0 else 0
                        break_even_roas = round((sell_price / margin_profit_unit) * 100, 2) if margin_profit_unit > 0 else 0

                        st.markdown(f"**판매가:** {sell_price:,}원")
                        st.markdown(f"**수수료율:** {fee_rate:.2f}%")
                        st.markdown(f"**매입비:** {product_data.get('purchase_cost', 0):,}원")
                        st.markdown(f"**수량:** {display_qty:,}개")
                        st.markdown(f"**매입단가:** {unit_purchase:,.0f}원")
                        st.markdown(f"**입출고/배송비:** {inout_shipping:,}원")
                        st.markdown(f"**물류비:** {product_data.get('logistics_cost', 0):,}원")
                        st.markdown(f"**관세:** {product_data.get('customs_duty', 0):,}원")
                        st.markdown(f"**기타:** {product_data.get('etc_cost', 0):,}원")
                        st.markdown(f"**마진율:** {margin_rate_pct:.2f}%")
                        st.markdown(f"**손익분기 ROAS:** {break_even_roas:.2f}%")
                    else:
                        st.info("선택된 상품의 상세 정보가 없습니다.")

                report_date = st.date_input("날짜 선택", datetime.date.today() - datetime.timedelta(days=1))
                st.markdown("---")

                st.markdown("#### 전체 판매")
                st.number_input("전체 판매 수량", step=1, key="total_sales_qty")
                st.number_input("전체 매출액", step=1000, key="total_revenue")
                st.number_input("개당 쿠폰가 (원)", step=100, key="coupon_unit")

                st.markdown("---")
                st.markdown("#### 광고 판매")
                st.number_input("광고 전환 판매 수량", step=1, key="ad_sales_qty")
                st.number_input("광고 매출액", step=1000, key="ad_revenue")
                st.number_input("광고비용", step=1000, key="ad_cost")

                st.markdown("---")
                st.markdown("#### 자연 판매 (자동 계산)")

                total_sales_qty = st.session_state.total_sales_qty
                display_revenue = st.session_state.total_revenue
                ad_sales_qty = st.session_state.ad_sales_qty
                ad_revenue_input = st.session_state.ad_revenue
                coupon_unit = st.session_state.get("coupon_unit", 0)

                coupon_total = coupon_unit * total_sales_qty
                actual_revenue = max(display_revenue - coupon_total, 0)

                ad_coupon_total = coupon_unit * ad_sales_qty
                ad_revenue_after_coupon = max(ad_revenue_input - ad_coupon_total, 0)

                organic_sales_qty_calc = max(total_sales_qty - ad_sales_qty, 0)
                organic_revenue_calc = max(actual_revenue - ad_revenue_after_coupon, 0)

                st.number_input("자연 판매 수량", value=organic_sales_qty_calc, disabled=True)
                st.number_input("자연 판매 매출액", value=organic_revenue_calc, disabled=True)

                if selected_product_name != "상품을 선택해주세요" and product_data:
                    current_total_sales_qty = st.session_state.total_sales_qty
                    display_revenue = st.session_state.total_revenue
                    current_ad_cost = st.session_state.ad_cost
                    coupon_unit = st.session_state.get("coupon_unit", 0)
                    coupon_total = coupon_unit * current_total_sales_qty
                    current_total_revenue = max(display_revenue - coupon_total, 0)

                    quantity_val = product_data.get("quantity", 1)
                    quantity_for_calc = quantity_val if quantity_val > 0 else 1
                    unit_purchase_cost = product_data.get("purchase_cost", 0) / quantity_for_calc
                    unit_logistics = product_data.get("logistics_cost", 0) / quantity_for_calc
                    unit_customs = product_data.get("customs_duty", 0) / quantity_for_calc
                    unit_etc = product_data.get("etc_cost", 0) / quantity_for_calc
                    fee_rate_db = product_data.get("fee", 0.0)

                    daily_profit = (
                        current_total_revenue
                        - (current_total_revenue * fee_rate_db / 100 * 1.1)
                        - (unit_purchase_cost * current_total_sales_qty)
                        - (product_data.get("inout_shipping_cost", 0) * current_total_sales_qty * 1.1)
                        - (unit_logistics * current_total_sales_qty)
                        - (unit_customs * current_total_sales_qty)
                        - (unit_etc * current_total_sales_qty)
                        - (current_ad_cost * 1.1)
                    )
                    daily_profit = won(daily_profit)
                    st.metric(label="일일 순이익금", value=f"{daily_profit:,}원")
                    # ✅ [복구] 일일 순이익 계산 내역(상세)
                    # (원본 65746에 있던 상세 출력 UI를 최신에 다시 추가)
                    vat = 1.1

                    fee_cost = won(current_total_revenue * fee_rate_db / 100 * vat)
                    purchase_cost_total = won(unit_purchase_cost * current_total_sales_qty)
                    inout_shipping_cost_total = won(product_data.get("inout_shipping_cost", 0) * current_total_sales_qty * vat)
                    logistics_cost_total = won(unit_logistics * current_total_sales_qty)
                    customs_cost_total = won(unit_customs * current_total_sales_qty)
                    etc_cost_total = won(unit_etc * current_total_sales_qty)
                    ad_cost_total = won(current_ad_cost * vat)

                    st.markdown(
                        f"""
                        <small>
                          - 판매 수수료 (VAT 포함): {format_number(fee_cost)}원<br>
                          - 상품 매입원가: {format_number(purchase_cost_total)}원<br>
                          - 입출고/배송비 (VAT 포함): {format_number(inout_shipping_cost_total)}원<br>
                          - 물류비: {format_number(logistics_cost_total)}원<br>
                          - 관세: {format_number(customs_cost_total)}원<br>
                          - 기타 비용: {format_number(etc_cost_total)}원<br>
                          - 광고비 (VAT 포함): {format_number(ad_cost_total)}원
                       </small>
                        """,
                        unsafe_allow_html=True
                    )

                st.markdown("---")

                if st.button("판매 기록 저장"):
                    if selected_product_name == "상품을 선택해주세요":
                        st.error("상품을 먼저 선택해야 판매 기록을 저장할 수 있습니다.")
                    elif not can_save_daily_record(
                        st.session_state.total_sales_qty,
                        st.session_state.total_revenue,
                        st.session_state.ad_sales_qty,
                        st.session_state.ad_revenue,
                        st.session_state.ad_cost,
                    ):
                        st.error("판매(수량/매출) 또는 광고(광고비/전환수/광고매출) 중 1개는 입력해야 저장할 수 있습니다.")

                    else:
                        try:
                            current_total_sales_qty = st.session_state.total_sales_qty
                            display_revenue = st.session_state.total_revenue
                            ad_sales_qty = st.session_state.ad_sales_qty
                            ad_revenue_input = st.session_state.ad_revenue
                            coupon_unit = st.session_state.get("coupon_unit", 0)

                            coupon_total = coupon_unit * current_total_sales_qty
                            current_total_revenue = max(display_revenue - coupon_total, 0)

                            ad_coupon_total = coupon_unit * ad_sales_qty
                            ad_revenue_after_coupon = max(ad_revenue_input - ad_coupon_total, 0)

                            organic_sales_qty = max(current_total_sales_qty - ad_sales_qty, 0)
                            organic_revenue = max(current_total_revenue - ad_revenue_after_coupon, 0)

                            quantity_val = product_data.get("quantity", 1) or 1
                            quantity_for_calc = quantity_val if quantity_val > 0 else 1

                            unit_purchase_cost = product_data.get("purchase_cost", 0) / quantity_for_calc
                            unit_logistics     = product_data.get("logistics_cost", 0) / quantity_for_calc
                            unit_customs       = product_data.get("customs_duty", 0) / quantity_for_calc
                            unit_etc           = product_data.get("etc_cost", 0) / quantity_for_calc

                            base_unit_cost = unit_purchase_cost + unit_logistics + unit_customs + unit_etc
                            invest_for_day = base_unit_cost * current_total_sales_qty
                            daily_roi = round(daily_profit / invest_for_day * 100, 2) if invest_for_day > 0 else 0

                            data_to_save = {
                                "date": report_date.isoformat(),
                                "product_name": selected_product_name,
                                "daily_sales_qty": current_total_sales_qty,
                                "daily_revenue": current_total_revenue,
                                "ad_sales_qty": ad_sales_qty,
                                "ad_revenue": ad_revenue_after_coupon,
                                "organic_sales_qty": organic_sales_qty,
                                "organic_revenue": organic_revenue,
                                "daily_ad_cost": st.session_state.ad_cost,
                                "daily_profit": daily_profit,
                                "daily_roi": daily_roi,
                            }

                            supabase.rpc("upsert_daily_sales", {"p_data": data_to_save}).execute()

                            st.success(
                                f"{report_date} 일일 판매 기록이 저장되었습니다! "
                                f"(순이익: {format_number(daily_profit)}원, ROI: {daily_roi}%)"
                            )

                        except Exception as e:
                            st.error(f"판매 기록 저장 중 오류가 발생했습니다: {e}")

    with tab4: # 원본 파일의 '세부 마진 계산기' 탭 내부의 '판매 현황' 내용
        c1, c2, c3, c4 = st.columns([0.1, 0.5, 1, 0.6])
        with c2: 
            # 각 기간 계산
            periods = {
                "오늘": get_date_range("today"),
                "어제": get_date_range("yesterday"),
                "7일": get_date_range("7days"),
                "15일": get_date_range("15days"),
                "30일": get_date_range("30days"),
                "90일": get_date_range("90days"),
                "180일": get_date_range("180days"),
                "365일": get_date_range("365days"),
            }

            for label, (start_d, end_d) in periods.items():
                profit_val = calculate_profit_for_period(start_d, end_d, supabase)
                st.markdown(
                    f"""
                    <div style='font-size:18px; margin-bottom:4px;'>
                        <span style='display:inline-block; width:50px; font-weight:bold;'>{label}</span>
                        <span style='display:inline-block; text-align:right; min-width:120px;'>{profit_val:,}원</span>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

        with c3:        
                st.markdown("#### 🗓️ 기간별 모든 상품 순이익 조회")

                today = datetime.date.today()
                last_7days_start, _ = get_date_range("7days")

                # 달력 2개 방식
                date_col1, date_col2 = st.columns(2)
                with date_col1:
                    start_date_input = st.date_input("시작 날짜", value=today, key="profit_start_date")
                with date_col2:
                    end_date_input = st.date_input("종료 날짜", value=today, key="profit_end_date")

                custom_profit = 0
                if start_date_input and end_date_input:
                    if start_date_input > end_date_input:
                        st.warning("시작 날짜는 종료 날짜보다 빠를 수 없습니다.")
                    else:
                        try:
                            custom_profit = calculate_profit_for_period(start_date_input, end_date_input, supabase)
                        except Exception as e:
                            st.error(f"지정 기간 순이익 계산 중 오류가 발생했습니다: {e}")

                st.metric(
                    label=f"선택 기간 ({start_date_input} ~ {end_date_input}) 모든 상품 총 순이익",
                    value=f"{format_number(custom_profit)}원"
                )


                # --- 페이지네이션 초기화 및 설정 --- (기존 코드 유지)
                def reset_page():
                    st.session_state.daily_sales_page = 1
                if 'daily_sales_page' not in st.session_state:
                    st.session_state.daily_sales_page = 1
                PAGE_SIZE = 20 # 한 페이지에 표시할 건수 (20건)

                # --- 상품 목록 로드 ---
                product_list = ["(상품을 선택해주세요)"]
                try:
                    response_prods = supabase.table("products").select("product_name").order("product_name").execute()
                    if response_prods.data:
                        product_list.extend([item['product_name'] for item in response_prods.data])
                except Exception as e:
                    st.warning("상품 목록을 불러올 수 없습니다. 상품 정보를 먼저 저장해주세요.")

                # --- 상품 필터 셀렉트 박스 ---
                selected_product_filter = st.selectbox(
                    "조회할 상품 선택",
                    product_list,
                    key="sales_status_product_filter",
                    on_change=reset_page # 필터 변경 시 페이지 1로 리셋
                )

                # 판매 현황 로직 시작
                try:
                    # 1. 데이터 로드 및 선택된 상품으로 필터링
                    query = supabase.table("daily_sales").select("*").order("date", desc=True)
            
                    # '상품을 선택해주세요'이 아닌 경우에만 쿼리에 필터 조건 추가
                    if selected_product_filter != "(상품을 선택해주세요)":
                        query = query.eq("product_name", selected_product_filter)
                
                    response = query.execute()
                    df = pd.DataFrame(response.data)

                    if not df.empty:
                        df['date'] = pd.to_datetime(df['date'])

                        # --- 특정 상품 선택 시에만 기록과 총 순이익금 표시 ---
                        if selected_product_filter != "(상품을 선택해주세요)":
                    
                            # 1. 총 합산 계산
                            total_profit_sum = df['daily_profit'].sum()
                            total_sales_qty = df['daily_sales_qty'].sum()
                            total_revenue_sum = df['daily_revenue'].sum()
                    
                            # 선택된 상품의 단가 정보를 로드 (ROI/마진율 계산에 필요)
                            product_data = {}
                            response_prod = supabase.table("products").select("*").eq("product_name", selected_product_filter).execute()
                            if response_prod.data:
                                product_data = response_prod.data[0]
                    
                            # 총 순이익금 표시
                            st.metric(label=f"총 순이익금 ({selected_product_filter})", value=f"{total_profit_sum:,}원")
                    
                            try:
                                # ROI / 마진율 계산에 필요한 총 수량과 단가 로드
                                total_quantity = product_data.get("quantity", 0) or 1
                                quantity_for_calc = total_quantity if total_quantity > 0 else 1
                                unit_purchase_cost = product_data.get("purchase_cost", 0) / quantity_for_calc
                                unit_logistics = product_data.get("logistics_cost", 0) / quantity_for_calc
                                unit_customs = product_data.get("customs_duty", 0) / quantity_for_calc
                                unit_etc = product_data.get("etc_cost", 0) / quantity_for_calc
                                inout_shipping_cost = product_data.get("inout_shipping_cost", 0)
                                fee_rate_db = product_data.get("fee", 0.0)

                                # ROI 분모 = 매입 + 물류 + 관세 + 기타 (총 순이익 블록과 동일)
                                purchase_cost_total = unit_purchase_cost * total_sales_qty
                                logistics_total = unit_logistics * total_sales_qty
                                customs_total = unit_customs * total_sales_qty
                                etc_total = unit_etc * total_sales_qty
                                total_cost_sum = purchase_cost_total + logistics_total + customs_total + etc_total # 이익이 아닌 총 원가/비용

                                # ROI / 마진율 계산 (총 순이익 블록)
                                roi = (total_profit_sum / total_cost_sum * 100) if total_cost_sum else 0
                                margin = (total_profit_sum / total_revenue_sum * 100) if total_revenue_sum else 0
                                                                
                                # 표시 블록 (세로 정렬)
                                st.markdown(
                                    f"""
                                    <div style='color:gray; font-size:14px; line-height:1.6;'>
                                        {total_quantity:,} / {total_sales_qty:,} (전체 수량 / 판매 수량)<br>
                                        ROI: {roi:.2f}%<br>
                                        마진율: {margin:.2f}%
                                    </div>
                                    """, unsafe_allow_html=True)
                            except Exception as e:
                                st.error(f"ROI/마진율 계산 중 오류 발생: {e}")
                    
                            st.markdown("---") # 순이익금과 기록 섹션 구분
                            st.markdown("#### 일일 판매 기록")

                        # 2. 페이지네이션 적용 로직
                        total_rows = len(df)
                        total_pages = (total_rows + PAGE_SIZE - 1) // PAGE_SIZE
                
                        if st.session_state.daily_sales_page > total_pages:
                            st.session_state.daily_sales_page = total_pages
                        if st.session_state.daily_sales_page < 1:
                            st.session_state.daily_sales_page = 1

                        start_index = (st.session_state.daily_sales_page - 1) * PAGE_SIZE
                        end_index = start_index + PAGE_SIZE
                
                        # 페이지에 맞는 데이터프레임 슬라이싱 (10일치)
                        df_paged = df.iloc[start_index:end_index].copy()

                        # 3. 컬럼명 변경 및 포맷팅
                        df_display = df_paged.rename(columns={
                            "date": "날짜",
                            "product_name": "상품명",
                            "daily_sales_qty": "판매량",
                            "daily_revenue": "매출액",
                            "ad_sales_qty": "광고 수량",
                            "ad_revenue": "광고 매출액",
                            "organic_sales_qty": "자연 수량",
                            "organic_revenue": "자연 매출액",
                            "daily_ad_cost": "광고비",
                            "daily_profit": "순이익",
                            # ROI / 마진율은 있으면 그대로 사용
                        })

                        # 날짜 포맷: 25-11-19 형식
                        df_display['날짜'] = df_display['날짜'].dt.strftime('%y-%m-%d')

                        # 최종 표시 컬럼 순서 (제목 줄여서 폭 확보)
                        display_cols = [
                            '날짜',
                            '상품명',
                            '판매량',
                            '매출액',
                            '광고 매출액',
                            '자연 매출액',
                            '광고비',
                            '순이익',
                        ]

                        # ── 계산용 원본 값들 (df_paged 기준, 숫자 그대로 사용) ──
                        sales_qty_vals       = df_paged["daily_sales_qty"].fillna(0)
                        total_revenue_vals   = df_paged["daily_revenue"].fillna(0)
                        ad_revenue_vals      = df_paged["ad_revenue"].fillna(0)
                        organic_revenue_vals = df_paged["organic_revenue"].fillna(0)
                        ad_cost_vals         = df_paged["daily_ad_cost"].fillna(0)
                        profit_vals          = df_paged["daily_profit"].fillna(0)

                        # 판매량: 쉼표만
                        df_display["판매량"] = sales_qty_vals.astype(int).apply(
                            lambda x: f"{x:,}"
                        )

                        # 매출액: 금액(원)만 콤마
                        df_display["매출액"] = total_revenue_vals.astype(int).apply(
                            lambda x: f"{x:,}"
                        )

                        # 광고 매출액: 금액 + 전체 매출 대비 비중
                        df_display["광고 매출액"] = [
                            f"{int(ad):,}({int(round(ad / tot * 100)) if tot > 0 else 0}%)"
                            for ad, tot in zip(ad_revenue_vals, total_revenue_vals)
                        ]

                        df_display["자연 매출액"] = [
                            f"{int(org):,}({int(round(org / tot * 100)) if tot > 0 else 0}%)"
                            for org, tot in zip(organic_revenue_vals, total_revenue_vals)
                        ]

                        df_display["광고비"] = [
                            f"{int(cost):,}({int(round(ad / cost * 100)) if cost > 0 else 0}%)"
                            for cost, ad in zip(ad_cost_vals, ad_revenue_vals)
                        ]

                        # ── 순이익 + ROI(%) 표시 ──
                        # daily_sales 테이블에 저장된 daily_roi 사용
                        if "daily_roi" in df_paged.columns:
                            roi_vals = df_paged["daily_roi"].fillna(0)
                            df_display["순이익"] = [
                                f"{int(p):,}({int(round(r))}%)"
                                for p, r in zip(profit_vals, roi_vals)
                            ]
                        else:
                            df_display["순이익"] = profit_vals.astype(int).apply(
                                lambda x: f"{x:,}"
                            )

                        # 4. 최종 데이터프레임 출력
                        st.dataframe(
                            df_display[display_cols],
                            hide_index=True,
                            use_container_width=True,  # 화면 가로 전체 사용
                            height=740,  # 이 줄 추가 (행 1개 약 35px + 헤더 35px)

                        )
                
                        # 5. 페이지네이션 버튼
                        page_cols = st.columns([1, 1, 1])
                        if page_cols[0].button("이전", disabled=(st.session_state.daily_sales_page <= 1), key="prev_page_btn"):
                            st.session_state.daily_sales_page -= 1
                            st.rerun()

                        page_cols[1].markdown(
                            f"<div style='text-align:center; font-size:16px; margin-top:5px;'>페이지 {st.session_state.daily_sales_page} / {total_pages}</div>", 
                            unsafe_allow_html=True
                        )

                        if page_cols[2].button("다음", disabled=(st.session_state.daily_sales_page >= total_pages), key="next_page_btn"):
                            st.session_state.daily_sales_page += 1
                            st.rerun()

                        st.markdown("---")

                    else:
                        st.info("아직 저장된 판매 기록이 없습니다.")
                except Exception as e:
                    st.error(f"판매 현황을 불러오는 중 오류가 발생했습니다: {e}")

                st.markdown("---")
                st.subheader("상품별 누적 매입 현황 (전체 차수 합산)")

                try:
                    # Supabase에서 매입 데이터 전체 가져오기
                    p_res = supabase.table("products").select("product_name, purchase_cost, logistics_cost, customs_duty").execute()
                    
                    if p_res.data:
                        df_p = pd.DataFrame(p_res.data)
                        
                        # 1. 'n차' 제거 및 이름 통일
                        df_p['rep_name'] = df_p['product_name'].apply(lambda x: re.sub(r'\d+차', '', str(x)).strip())
                        
                        # 2. 상품별 합산 (매입비, 물류비, 관세)
                        p_summary = df_p.groupby('rep_name').agg({
                            'purchase_cost': 'sum',
                            'logistics_cost': 'sum',
                            'customs_duty': 'sum'
                        }).reset_index()
                        
                        # 3. 상품별 총 합계 열 추가
                        p_summary['item_total'] = p_summary['purchase_cost'] + p_summary['logistics_cost'] + p_summary['customs_duty']
                        p_summary = p_summary.sort_values('rep_name')

                        # 4. 전체 총 합계 행(Total Row) 계산
                        total_row = pd.DataFrame([{
                            'rep_name': '총 합계',
                            'purchase_cost': p_summary['purchase_cost'].sum(),
                            'logistics_cost': p_summary['logistics_cost'].sum(),
                            'customs_duty': p_summary['customs_duty'].sum(),
                            'item_total': p_summary['item_total'].sum()
                        }])
                        
                        # 결과 합치기
                        final_df = pd.concat([p_summary, total_row], ignore_index=True)

                        # 5. 천 단위 콤마 포맷팅
                        formatted_df = final_df.copy()
                        for col in ['purchase_cost', 'logistics_cost', 'customs_duty', 'item_total']:
                            formatted_df[col] = formatted_df[col].apply(lambda x: f"{int(x):,}")

                        # 컬럼명 변경 후 출력
                        formatted_df.columns = ["대표 상품명", "총 매입비", "총 물류비", "총 관세", "상품별 총 합계"]
                        st.dataframe(formatted_df, hide_index=True, use_container_width=True)
                    else:
                        st.info("등록된 매입 데이터가 없습니다.")
                except Exception as e:
                    st.error(f"누적 매입 현황 계산 중 오류: {e}")

    with tab5:
        render_ad_analysis_tab(supabase)

if __name__ == "__main__":
    # 메인 실행 전에 탭 1의 세션 상태 키 초기화 보장
    if "sell_price_raw" not in st.session_state: st.session_state["sell_price_raw"] = ""
    if "unit_yuan" not in st.session_state: st.session_state["unit_yuan"] = ""
    if "unit_won" not in st.session_state: st.session_state["unit_won"] = ""
    if "qty_raw" not in st.session_state: st.session_state["qty_raw"] = ""
    main()
